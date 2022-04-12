# Q3 : Find max amount and min amount in the output
from openpyxl import load_workbook


empty_amount = []
#validating the max and min amount from deposits sheet here:
wb = load_workbook("task_output.xlsx")
ws = wb['DEPOSITS']
column = ws['C']
column_list = [column[x].value for x in range(len(column))]
triming_header = column_list[1:]
empty_amount.append(max(triming_header))
empty_amount.append(min(triming_header))


#validating the max and min amount from withdrawl sheet here:
wb = load_workbook("task_output.xlsx")
ws = wb['WITHDRAWALS']
column_wd = ws['C']
column_list_wd = [column_wd[x].value for x in range(len(column_wd))]
triming_header_wd = column_list_wd[1:]
empty_amount.append(max(triming_header_wd))
empty_amount.append(min(triming_header_wd))

maximum = max(empty_amount)
minimum = min(empty_amount)

print("Maximum amount:",maximum)
print("minimum amount:",minimum)