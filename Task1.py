import json


empty = []
#Accessing the json and exposing data:
json_file = open('task_input_list.json')
data_parsing = json.load(json_file)
for i in data_parsing:
   empty.append(i)



#Perform validations on date and amount columns

from openpyxl import load_workbook
import datetime

#validating the date here:
wb = load_workbook("task_output.xlsx")
ws = wb['DEPOSITS']
column = ws['A']
column_list = [column[x].value for x in range(len(column))]
date_trim = column_list[1:]
for inputDate in date_trim:
    day, month, year = inputDate.split('/')
    isValidDate = True
    try:
        datetime.datetime(int(year), int(month), int(day))
    except ValueError:
        isValidDate = False
    if(isValidDate):
        print("Input date is valid ..",inputDate)
    else:
        print("Input date is not valid..",inputDate)


#validating the amount here:
wb = load_workbook("task_output.xlsx")
ws = wb['DEPOSITS']
column = ws['C']
column_list = [column[x].value for x in range(len(column))]
amount_trim = column_list[1:]
for j in amount_trim:
    if(isinstance(j, int)):
        print("amount is valid",j)
    else:
        print("amount is In-valid",j)



